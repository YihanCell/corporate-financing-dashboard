from __future__ import annotations

import json
import math
import os
import re
import tempfile
from datetime import date, datetime
from io import BytesIO
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote, unquote, urlparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
UPLOAD_DIR = ROOT / "uploads"
EXPORT_DIR = ROOT / "data"
PAYLOAD_PATH = EXPORT_DIR / "current_payload.json"
FALLBACK_PAYLOAD_PATH = Path(r"C:\tmp\finance-dashboard\current_payload.json")
DEFAULT_SOURCE_DIR = Path(
    r"C:\Users\Administrator\Documents\xwechat_files\Oscar19_bb53\temp\RWTemp\2026-05\ec86b6af4ed5d2f92c94c55cff820f25"
)

COL = {
    "index": "\u5e8f\u53f7",
    "borrower": "\u878d\u8d44\u4e3b\u4f53",
    "institution": "\u673a\u6784\u540d\u79f0",
    "product": "\u878d\u8d44\u54c1\u79cd",
    "purpose": "\u7528\u9014",
    "credit": "\u6388\u4fe1\u989d\u5ea6\uff08\u4e07\u5143\uff09",
    "draw": "\u63d0\u6b3e\u91d1\u989d\uff08\u4e07\u5143\uff09",
    "balance": "\u8d37\u6b3e\u4f59\u989d\uff08\u4e07\u5143\uff09",
    "rate": "\u5f53\u524d\u5229\u7387",
    "contract_rate": "\u5408\u540c\u5229\u7387",
    "start_date": "\u8d77\u606f\u65e5\u671f",
    "maturity_date": "\u5230\u671f\u65e5\u671f",
    "term": "\u8fd8\u6b3e\u671f\u9650",
    "guarantee": "\u62c5\u4fdd\u65b9\u5f0f",
    "repayment_plan": "\u8fd8\u6b3e\u8ba1\u5212",
    "notes": "\u5907\u6ce8",
}

LATEST_PAYLOAD: dict | None = None


def save_current_payload(payload: dict) -> None:
    data = json.dumps(payload, ensure_ascii=False)
    for path in (PAYLOAD_PATH, FALLBACK_PAYLOAD_PATH):
        try:
            path.parent.mkdir(exist_ok=True)
            path.write_text(data, encoding="utf-8")
            return
        except OSError:
            continue


def load_current_payload() -> dict | None:
    for path in (PAYLOAD_PATH, FALLBACK_PAYLOAD_PATH):
        if not path.exists():
            continue
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
    return None


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if hasattr(value, "item"):
        return clean_value(value.item())
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def to_number(value):
    value = clean_value(value)
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("\uff0c", "").replace("%", "").strip()
    if not text:
        return 0.0
    try:
        number = float(text)
    except ValueError:
        return 0.0
    if "%" in str(value):
        number = number / 100
    return number


def to_date(value):
    value = clean_value(value)
    if value is None:
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date().isoformat()


def first_existing_excel(folder: Path) -> Path | None:
    files = sorted(folder.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def parse_workbook(source, source_name: str | None = None) -> dict:
    if isinstance(source, (bytes, bytearray)):
        if len(source) > 80 * 1024 * 1024:
            raise ValueError("Excel 文件过大，请确认上传的是融资情况表。")
        workbook_source = BytesIO(source)
        display_name = source_name or "已上传融资情况表.xlsx"
    else:
        path = Path(source)
        if path.stat().st_size > 80 * 1024 * 1024:
            raise ValueError("Excel 文件过大，请确认上传的是融资情况表。")
        workbook_source = path
        display_name = path.name

    xls = pd.ExcelFile(workbook_source)
    if len(xls.sheet_names) < 2:
        raise ValueError("Excel \u81f3\u5c11\u9700\u8981\u4e24\u4e2a\u5de5\u4f5c\u8868\u3002")

    sheet_name = xls.sheet_names[1]
    header = pd.read_excel(xls, sheet_name=sheet_name, nrows=0)
    required = {COL["borrower"], COL["institution"], COL["product"], COL["balance"], COL["maturity_date"]}
    missing = [name for name in required if name not in set(map(str, header.columns))]
    if missing:
        raise ValueError("第二个工作表不像融资台账，缺少字段：" + "、".join(missing))

    frame = pd.read_excel(xls, sheet_name=sheet_name)
    frame = frame.dropna(how="all")

    records = []
    carry = {
        COL["borrower"]: None,
        COL["institution"]: None,
        COL["product"]: None,
        COL["purpose"]: None,
    }

    for _, row in frame.iterrows():
        item = {str(col): clean_value(row[col]) for col in frame.columns}
        original_item = item.copy()
        if not item.get(COL["index"]) and not item.get(COL["balance"]):
            continue

        has_original_identity = any(
            original_item.get(key)
            for key in (COL["index"], COL["borrower"], COL["institution"], COL["product"], COL["purpose"])
        )
        has_original_timing = any(
            original_item.get(key) for key in (COL["start_date"], COL["maturity_date"], COL["term"])
        )
        if not has_original_identity and not has_original_timing:
            continue

        for key in carry:
            if item.get(key):
                carry[key] = item[key]
            else:
                item[key] = carry[key]

        balance = to_number(item.get(COL["balance"]))
        credit = to_number(item.get(COL["credit"]))
        draw = to_number(item.get(COL["draw"]))
        rate = to_number(item.get(COL["rate"]))
        start_date = to_date(item.get(COL["start_date"]))
        maturity_date = to_date(item.get(COL["maturity_date"]))

        if balance <= 0 and credit <= 0 and draw <= 0:
            continue

        records.append(
            {
                "index": clean_value(item.get(COL["index"])),
                "borrower": item.get(COL["borrower"]) or "\u672a\u586b\u5199",
                "institution": item.get(COL["institution"]) or "\u672a\u586b\u5199",
                "product": item.get(COL["product"]) or "\u672a\u586b\u5199",
                "purpose": item.get(COL["purpose"]) or "",
                "credit": credit,
                "draw": draw,
                "balance": balance,
                "rate": rate,
                "contractRate": clean_value(item.get(COL["contract_rate"])),
                "startDate": start_date,
                "maturityDate": maturity_date,
                "term": clean_value(item.get(COL["term"])),
                "guarantee": clean_value(item.get(COL["guarantee"])),
                "repaymentPlan": clean_value(item.get(COL["repayment_plan"])),
                "notes": clean_value(item.get(COL["notes"])),
            }
        )

    active_records = [record for record in records if record["balance"] > 0]
    return {
        "sourceFile": display_name,
        "sheetName": sheet_name,
        "loadedAt": datetime.now().isoformat(timespec="seconds"),
        "rows": active_records,
        "rowCount": len(active_records),
        "allRowCount": len(records),
        "sheetNames": xls.sheet_names,
    }


def is_external_detail_row(row: dict) -> bool:
    product = str(row.get("product") or "")
    excluded_terms = ("\u516c\u53f8\u503a", "\u4f01\u4e1a\u503a", "\u503a\u5238", "\u4e2d\u671f\u7968\u636e")
    return row.get("balance", 0) > 0 and not any(term in product for term in excluded_terms)


def create_external_detail_workbook(payload: dict) -> bytes:
    rows = [row for row in payload.get("rows", []) if is_external_detail_row(row)]

    wb = Workbook()
    ws = wb.active
    ws.title = "\u878d\u8d44\u660e\u7ec6"
    headers = [
        "\u5e8f\u53f7",
        "\u878d\u8d44\u4e3b\u4f53",
        "\u673a\u6784\u540d\u79f0",
        "\u878d\u8d44\u54c1\u79cd",
        "\u8d37\u6b3e\u4f59\u989d\uff08\u4e07\u5143\uff09",
        "\u5230\u671f\u65e5\u671f",
    ]
    ws.append(headers)

    for idx, row in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                row.get("borrower"),
                row.get("institution"),
                row.get("product"),
                row.get("balance"),
                row.get("maturityDate"),
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border_side = Side(style="thin", color="9AA6B2")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in enumerate([8, 34, 24, 18, 18, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for cell in ws["E"][1:]:
        cell.number_format = '#,##0.00'
    for cell in ws["F"][1:]:
        cell.number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A2"

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as handle:
        temp_path = Path(handle.name)
    wb.save(temp_path)
    data = temp_path.read_bytes()
    temp_path.unlink(missing_ok=True)
    return data


def create_filtered_detail_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "\u7b5b\u9009\u660e\u7ec6"
    headers = [
        "\u5e8f\u53f7",
        "\u878d\u8d44\u4e3b\u4f53",
        "\u673a\u6784\u540d\u79f0",
        "\u878d\u8d44\u54c1\u79cd",
        "\u7528\u9014",
        "\u6388\u4fe1\u989d\u5ea6\uff08\u4e07\u5143\uff09",
        "\u63d0\u6b3e\u91d1\u989d\uff08\u4e07\u5143\uff09",
        "\u8d37\u6b3e\u4f59\u989d\uff08\u4e07\u5143\uff09",
        "\u5f53\u524d\u5229\u7387",
        "\u8d77\u606f\u65e5\u671f",
        "\u5230\u671f\u65e5\u671f",
        "\u8fd8\u6b3e\u671f\u9650",
        "\u62c5\u4fdd\u65b9\u5f0f",
        "\u8fd8\u6b3e\u8ba1\u5212",
        "\u5907\u6ce8",
    ]
    ws.append(headers)
    for idx, row in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                row.get("borrower"),
                row.get("institution"),
                row.get("product"),
                row.get("purpose"),
                row.get("credit"),
                row.get("draw"),
                row.get("balance"),
                row.get("rate"),
                row.get("startDate"),
                row.get("maturityDate"),
                row.get("term"),
                row.get("guarantee"),
                row.get("repaymentPlan"),
                row.get("notes"),
            ]
        )

    if rows:
        summary_row = ws.max_row + 1
        ws.cell(summary_row, 1, "\u5408\u8ba1")
        ws.cell(summary_row, 8, sum(to_number(row.get("balance")) for row in rows))
        denominator = sum(to_number(row.get("balance")) for row in rows)
        if denominator:
            ws.cell(
                summary_row,
                9,
                sum(to_number(row.get("balance")) * to_number(row.get("rate")) for row in rows) / denominator,
            )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="102B46")
    border_side = Side(style="thin", color="9AA6B2")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif cell.row == ws.max_row and rows:
                cell.fill = total_fill
                cell.font = Font(bold=True, color="FFFFFF")
    for col, width in enumerate([8, 34, 24, 18, 28, 16, 16, 18, 12, 14, 14, 14, 28, 30, 36], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for cell in ws["F"][1:] + ws["G"][1:] + ws["H"][1:]:
        cell.number_format = '#,##0.00'
    for cell in ws["I"][1:]:
        cell.number_format = "0.00%"
    ws.freeze_panes = "A2"

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as handle:
        temp_path = Path(handle.name)
    wb.save(temp_path)
    data = temp_path.read_bytes()
    temp_path.unlink(missing_ok=True)
    return data


class DashboardHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            self.send_json({"ok": True})
            return
        if parsed.path == "/api/current":
            self.respond_with_current()
            return
        if parsed.path == "/api/load-default":
            default_file = first_existing_excel(DEFAULT_SOURCE_DIR)
            if not default_file:
                self.send_json({"error": "\u6ca1\u6709\u627e\u5230\u9ed8\u8ba4 Excel \u6587\u4ef6\u3002"}, HTTPStatus.NOT_FOUND)
                return
            self.respond_with_workbook(default_file)
            return
        if parsed.path == "/api/export-details":
            self.respond_with_export()
            return

        requested = unquote(parsed.path.lstrip("/")) or "index.html"
        target = (STATIC_DIR / requested).resolve()
        if not str(target).startswith(str(STATIC_DIR.resolve())):
            self.send_error(HTTPStatus.FORBIDDEN)
            return
        if target.is_dir():
            target = target / "index.html"
        if not target.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        self.path = "/" + target.relative_to(STATIC_DIR).as_posix()
        return SimpleHTTPRequestHandler.do_GET(self)

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/api/export-filtered":
            self.respond_with_filtered_export()
            return
        if path != "/api/upload":
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_type = self.headers.get("Content-Type", "")
        length = int(self.headers.get("Content-Length", "0"))
        if not content_type.lower().startswith("multipart/form-data"):
            self.send_json({"error": "\u8bf7\u4f7f\u7528\u8868\u5355\u4e0a\u4f20 Excel \u6587\u4ef6\u3002"}, HTTPStatus.BAD_REQUEST)
            return
        if length <= 0:
            self.send_json({"error": "\u6ca1\u6709\u6536\u5230 Excel \u6587\u4ef6\u3002"}, HTTPStatus.BAD_REQUEST)
            return
        if length > 80 * 1024 * 1024:
            self.send_json({"error": "Excel \u6587\u4ef6\u8fc7\u5927\uff0c\u8bf7\u786e\u8ba4\u4e0a\u4f20\u7684\u662f\u878d\u8d44\u60c5\u51b5\u8868\u3002"}, HTTPStatus.BAD_REQUEST)
            return

        self.connection.settimeout(8)
        chunks = []
        remaining = length
        try:
            while remaining > 0:
                chunk = self.rfile.read(min(65536, remaining))
                if not chunk:
                    break
                chunks.append(chunk)
                remaining -= len(chunk)
        except TimeoutError:
            pass
        body = b"".join(chunks)
        if not body:
            self.send_json({"error": "\u6ca1\u6709\u8bfb\u5230 Excel \u6587\u4ef6\u5185\u5bb9\u3002"}, HTTPStatus.BAD_REQUEST)
            return
        boundary_match = re.search(r"boundary=(.+)", content_type)
        if not boundary_match:
            self.send_json({"error": "\u6ca1\u6709\u6536\u5230 Excel \u6587\u4ef6\u3002"}, HTTPStatus.BAD_REQUEST)
            return
        boundary = boundary_match.group(1).strip().strip('"').encode("utf-8")
        file_bytes = None
        filename = None
        for part in body.split(b"--" + boundary):
            if b"Content-Disposition:" not in part or b"filename=" not in part:
                continue
            header_blob, _, content = part.partition(b"\r\n\r\n")
            if not content:
                continue
            disposition = header_blob.decode("utf-8", errors="replace")
            name_match = re.search(r'filename="([^"]*)"', disposition)
            filename = Path(name_match.group(1)).name if name_match else "upload.xlsx"
            file_bytes = content[:-2] if content.endswith(b"\r\n") else content
            break
        if not filename or file_bytes is None:
            self.send_json({"error": "\u6ca1\u6709\u6536\u5230 Excel \u6587\u4ef6\u3002"}, HTTPStatus.BAD_REQUEST)
            return

        if not filename.lower().endswith(".xlsx"):
            self.send_json({"error": "\u8bf7\u4e0a\u4f20 .xlsx \u6587\u4ef6\u3002"}, HTTPStatus.BAD_REQUEST)
            return
        self.respond_with_workbook(file_bytes, filename)

    def read_json_body(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def respond_with_current(self):
        global LATEST_PAYLOAD
        if LATEST_PAYLOAD is None:
            LATEST_PAYLOAD = load_current_payload()
        if not LATEST_PAYLOAD:
            self.send_json({"error": "\u670d\u52a1\u7aef\u5c1a\u672a\u4e0a\u4f20\u878d\u8d44\u60c5\u51b5\u8868\u3002"}, HTTPStatus.NOT_FOUND)
            return
        self.send_json(LATEST_PAYLOAD)

    def respond_with_workbook(self, source, source_name: str | None = None):
        global LATEST_PAYLOAD
        try:
            payload = parse_workbook(source, source_name)
        except Exception as exc:
            self.send_json({"error": f"\u8bfb\u53d6\u5931\u8d25\uff1a{exc}"}, HTTPStatus.BAD_REQUEST)
            return
        LATEST_PAYLOAD = payload
        save_current_payload(payload)
        self.send_json(payload)

    def respond_with_export(self):
        if not LATEST_PAYLOAD:
            self.send_json({"error": "\u8bf7\u5148\u8bfb\u53d6\u4e00\u4efd\u878d\u8d44\u60c5\u51b5\u8868\u3002"}, HTTPStatus.BAD_REQUEST)
            return
        data = create_external_detail_workbook(LATEST_PAYLOAD)
        filename = f"\u878d\u8d44\u660e\u7ec6\uff08{datetime.now():%Y%m%d}\uff09.xlsx"
        encoded = quote(filename)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{encoded}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def respond_with_filtered_export(self):
        try:
            payload = self.read_json_body()
            rows = payload.get("rows", [])
            if not isinstance(rows, list):
                raise ValueError("\u7b5b\u9009\u660e\u7ec6\u683c\u5f0f\u4e0d\u6b63\u786e\u3002")
        except Exception as exc:
            self.send_json({"error": f"\u5bfc\u51fa\u5931\u8d25\uff1a{exc}"}, HTTPStatus.BAD_REQUEST)
            return
        data = create_filtered_detail_workbook(rows)
        filename = f"\u7b5b\u9009\u878d\u8d44\u660e\u7ec6\uff08{datetime.now():%Y%m%d}\uff09.xlsx"
        encoded = quote(filename)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{encoded}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload, status=HTTPStatus.OK):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def main():
    global LATEST_PAYLOAD
    LATEST_PAYLOAD = load_current_payload()
    host = os.environ.get("FINANCE_DASHBOARD_HOST", "127.0.0.1")
    port = int(os.environ.get("FINANCE_DASHBOARD_PORT", "8780"))
    server = ThreadingHTTPServer((host, port), DashboardHandler)
    shown_host = "127.0.0.1" if host in ("", "0.0.0.0") else host
    print(f"\u878d\u8d44\u770b\u76d8\u5df2\u542f\u52a8\uff1ahttp://{shown_host}:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()

