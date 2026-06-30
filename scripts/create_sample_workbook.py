from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "samples" / "企业融资情况表示例.xlsx"


def main() -> None:
    OUT.parent.mkdir(exist_ok=True)

    wb = Workbook()
    ws_lpr = wb.active
    ws_lpr.title = "LPR"
    ws_lpr.append(["日期", "1年期LPR", "5年期以上LPR"])
    ws_lpr.append(["2026-06-20", 0.0300, 0.0350])
    ws_lpr.append(["说明", "示例数据", "非真实融资信息"])

    ws = wb.create_sheet("截至2026年6月30日")
    headers = [
        "序号",
        "融资主体",
        "机构名称",
        "融资品种",
        "用途",
        "授信额度（万元）",
        "提款金额（万元）",
        "贷款余额（万元）",
        "当前利率",
        "合同利率",
        "起息日期",
        "到期日期",
        "还款期限",
        "担保方式",
        "还款计划",
        "备注",
        "公盘链接",
    ]
    ws.append(headers)
    rows = [
        [1, "示例集团有限公司", "国家开发银行", "固定资产贷款", "产业园基础设施建设", 80000, 80000, 65000, 0.0320, 0.0320, "2024-01-15", "2029-01-15", "5年", "信用", "按季付息，到期还本", "示例数据", ""],
        [2, "示例集团有限公司", "工商银行", "流贷", "补充流动资金", 30000, 30000, 28000, 0.0265, 0.0265, "2026-05-20", "2026-07-05", "46天", "保证", "到期一次性还本付息", "7天内/30天内示例", ""],
        [3, "示例城市建设有限公司", "建设银行", "流贷", "项目周转", 25000, 25000, 18000, 0.0280, 0.0280, "2026-04-25", "2026-07-24", "90天", "信用", "到期一次性还本付息", "8-30天示例", ""],
        [4, "示例城市建设有限公司", "农业银行", "固定资产贷款", "停车场改造", 42000, 42000, 36000, 0.0310, 0.0310, "2024-08-01", "2026-08-20", "2年", "抵押", "按半年还本，按季付息", "31-60天示例", ""],
        [5, "示例投资发展有限公司", "交通银行", "流贷", "日常经营周转", 15000, 15000, 12000, 0.0245, 0.0245, "2026-03-01", "2026-09-18", "201天", "信用", "到期还本", "61-90天示例", ""],
        [6, "示例投资发展有限公司", "浦发银行", "项目贷款", "市政配套项目", 50000, 50000, 45500, 0.0345, 0.0345, "2025-02-10", "2026-11-15", "21个月", "保证", "分期还本", "91-180天示例", ""],
        [7, "示例资产运营有限公司", "中信银行", "流贷", "物业运营", 12000, 12000, 11800, 0.0215, 0.0215, "2026-01-10", "2027-03-10", "14个月", "信用", "到期还本", "181-365天示例", ""],
        [8, "示例资产运营有限公司", "招商银行", "固定资产贷款", "厂房改造", 36000, 36000, 30000, 0.0295, 0.0295, "2025-06-01", "2030-06-01", "5年", "抵押", "按季付息，分期还本", "三年以上示例", ""],
        [9, "示例集团有限公司", "证券公司A", "公司债", "偿还有息债务", 100000, 100000, 75000, 0.0330, 0.0330, "2025-09-01", "2028-09-01", "3年", "/", "每年付息，到期还本", "债券示例", ""],
        [10, "示例集团有限公司", "承销商B", "中期票据", "补充流动资金", 60000, 60000, 60000, 0.0305, 0.0305, "2026-01-20", "2029-01-20", "3年", "/", "每年付息，到期还本", "中期票据示例", ""],
        [11, "示例文旅发展有限公司", "兴业银行", "银团贷款", "文旅项目建设", 55000, 55000, 52000, 0.0360, 0.0360, "2023-12-01", "2031-12-01", "8年", "保证", "分期还本", "高成本长期贷款示例", ""],
        [12, "示例文旅发展有限公司", "民生银行", "流贷", "票据置换", 9000, 9000, 8800, 0.0190, 0.0190, "2026-06-01", "2026-07-02", "31天", "信用", "到期还本", "低利率示例", ""],
    ]
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border_side = Side(style="thin", color="AAB7C4")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [8, 28, 18, 16, 24, 16, 16, 16, 12, 12, 14, 14, 12, 16, 26, 24, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 9).number_format = "0.00%"
        ws.cell(row, 10).number_format = "0.00%"
        for col in (6, 7, 8):
            ws.cell(row, col).number_format = "#,##0.00"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
